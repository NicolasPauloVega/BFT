# Messages and errors validation
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q

# User and models
from .forms import *
from user.models import *

# Encript password and validation
from django.contrib.auth.hashers import make_password, check_password

# Email reset
from django.core.mail import send_mail

# User validate
from .middleware import *

# Chart
from django.db.models import Sum, Max, Min
from django.db.models.functions import TruncMonth
from datetime import datetime, date, timedelta

# Exports excel
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font

# Create your views here.
@login_required
@rol_required([1]) # Solo puede entrar un administrador
def index(request):
    count_user = Usuario.objects.count()
    count_evidence = Puntos.objects.count()
    programs = Categoria.objects.count()
    return render(request, "admin_dashboard/index.html", {'count': count_user, 'evidencia': count_evidence, 'programas_count': programs})

@login_required
@rol_required([1])  # Solo puede entrar un administrador
def charts(request):
    # Obtener los valores máximo y mínimo de `total_puntos`
    mejor_puntaje = Usuario.objects.aggregate(Max('total_puntos'))['total_puntos__max']
    peor_puntaje = Usuario.objects.aggregate(Min('total_puntos'))['total_puntos__min']

    # Asegurarse de que no haya errores si no hay datos
    top_usuarios = Usuario.objects.filter(total_puntos=mejor_puntaje) if mejor_puntaje else []
    worse_usuarios = Usuario.objects.filter(total_puntos=peor_puntaje) if peor_puntaje else []

    # Crear listas de nombres y puntos
    nombres = [usuario.nombre for usuario in top_usuarios]
    puntos = [usuario.total_puntos for usuario in top_usuarios]

    nombre_peor = [usuario.nombre for usuario in worse_usuarios]
    peor_puntos = [usuario.total_puntos for usuario in worse_usuarios]

    # Calcular puntuaciones mensuales y anuales
    mes_actual = datetime.now().month
    año_actual = datetime.now().year

    # Agrupar por usuario y calcular totales de puntos por mes
    puntos_mes = (
        Puntos.objects.filter(fecha__month=mes_actual, fecha__year=año_actual)
        .values('usuario')
        .annotate(total=Sum('puntos'))
    )

    puntos_anual = (
        Puntos.objects.filter(fecha__year=año_actual)
        .values('usuario')
        .annotate(total=Sum('puntos'))
    )

    # Mejores y peores puntuados del mes
    mejor_mes = puntos_mes.order_by('-total').first()
    peor_mes = puntos_mes.order_by('total').first()

    # Mejores y peores puntuados del año
    mejor_anual = puntos_anual.order_by('-total').first()
    peor_anual = puntos_anual.order_by('total').first()

    # Contar cuántos usuarios tienen esas puntuaciones
    count_mejor_mes = puntos_mes.filter(total=mejor_mes['total']).count() if mejor_mes else 0
    count_peor_mes = puntos_mes.filter(total=peor_mes['total']).count() if peor_mes else 0
    count_mejor_anual = puntos_anual.filter(total=mejor_anual['total']).count() if mejor_anual else 0
    count_peor_anual = puntos_anual.filter(total=peor_anual['total']).count() if peor_anual else 0

    # Enviar contexto adicional para gráficos
    context = {
        'nombres': nombres,
        'puntos': puntos,
        'peor_nombre': nombre_peor,
        'peor_puntaje': peor_puntos,
        'datos_mes': [count_mejor_mes, count_peor_mes],
        'datos_anual': [count_mejor_anual, count_peor_anual],
    }
    return render(request, "admin_dashboard/charts.html", context)

@login_required
@rol_required([1]) # Solo puede entrar un administrador
def tables(request):
    usuario = Usuario.objects.all()
    return render(request, "admin_dashboard/tables.html", {'user': usuario})

@login_required
@rol_required([1])
def category(request):
    categorias = Categoria.objects.all()
    categorias_data = [
        {
            'id': categoria.id,
            'nombre': categoria.nombre,
            'subcategorias': list(categoria.subcategorias.values('id', 'nombre', 'puntos_positivos', 'puntos_negativos')),
        }
        for categoria in categorias
    ]
    return render(request, "admin_dashboard/category/category.html", {'categoria': categorias_data})

@login_required
@rol_required([1])
def add_category(request):
    if request.method == 'POST':
        categoria = request.POST['nombre_categoria']
        
        if Categoria.objects.filter(nombre=categoria).exists(): # Validamos si la categoria exite
            messages.error(request, '¡La categoría ya existe!')
            return redirect('category')
        else:
            Categoria.objects.create(nombre=categoria) # Guardando la categoria
            messages.success(request, '¡Categoría creada exitosamente!')
            return redirect('category-subcategory')
    return render(request, "admin_dashboard/category/add_category.html")

@login_required
@rol_required([1])
def add_subcategory(request, categoria_id):
    # Obtener la categoría como instancia
    category = get_object_or_404(Categoria, id=categoria_id)
    
    if request.method == 'POST':
        # Recoger datos del formulario
        subcategoria = request.POST['nombre_subcategoria']
        puntos_positivos = request.POST['puntos_positivos']
        puntos_negativos = request.POST['puntos_negativos']
        
        # Verificar si la subcategoría ya existe
        if Subcategoria.objects.filter(nombre=subcategoria, categoria=category).exists():
            messages.error(request, '¡Esta subcategoría ya existe para esta categoría!')
            return redirect('subcategory', categoria_id=categoria_id)
        else:
            # Crear la subcategoría
            Subcategoria.objects.create(
                categoria=category,  # Usar la instancia de Categoria
                nombre=subcategoria,
                puntos_positivos=puntos_positivos,
                puntos_negativos=puntos_negativos
            )
            messages.success(request, '¡Subcategoría creada exitosamente!')
            return redirect('subcategory', categoria_id=categoria_id)
    
    return render(request, "admin_dashboard/category/add_subcategory.html", {'category': category})

@login_required
@rol_required([1])
def subcategory(request, categoria_id):
    subcategory=Subcategoria.objects.filter(categoria=categoria_id)
    return render(request, 'admin_dashboard/category/subcategory.html', {'subcategory': subcategory}) 

@login_required
@rol_required([1])
def points(request, id_usuario):
    # Obtener el usuario mediante una url
    usuario = get_object_or_404(Usuario, id=id_usuario)
    
    # Obtener todas las categorias
    categoria = Categoria.objects.all()
    subcategoria = Subcategoria.objects.all()
    
    return render(request, 'admin_dashboard/points.html', {
        'usuario': usuario,
        'categoria': categoria,
        'subcategorias':subcategoria,
    })

@login_required
@rol_required([1])
def save_form_points(request):
    if request.method == 'POST':
        # Obtener valores del formulario
        usuario_id = request.POST.get('idUsuario')  # ID del usuario
        subcategoria_id = request.POST.get('subcategoria')  # Subcategoría seleccionada
        evidencia = request.FILES.get('img')  # Archivo de evidencia (opcional)
        puntos = int(request.POST.get('puntos'))  # Valor enviado desde el input:radio

        # Obtener usuario
        try:
            usuario = Usuario.objects.get(id=usuario_id)
        except Usuario.DoesNotExist:
            messages.error(request, 'Usuario no encontrado')
            return redirect('points', id_usuario=usuario_id)

        # Obtener subcategoría y validar puntos
        try:
            subcategoria = Subcategoria.objects.get(id=subcategoria_id)
        except Subcategoria.DoesNotExist:
            messages.error(request, 'Subcategoría no encontrada')
            return redirect('points', id_usuario=usuario_id)

        if puntos == subcategoria.puntos_positivos:
            # Sumar puntos
            usuario.total_puntos += puntos
            puntos_registrados = puntos

        elif puntos == subcategoria.puntos_negativos:
            # Validar si puede restar los puntos
            if usuario.total_puntos < puntos:
                messages.error(request, 'El usuario no tiene suficientes puntos para restar')
                return redirect('points', id_usuario=usuario_id)

            # Restar puntos
            usuario.total_puntos -= puntos
            puntos_registrados = -puntos

        else:
            messages.error(request, 'Los puntos no coinciden con los valores permitidos en la subcategoría seleccionada')
            return redirect('points', id_usuario=usuario_id)

        # Guardar cambios
        usuario.save()
        if evidencia:
            Puntos.objects.create(
                usuario=usuario,
                subcategoria=subcategoria,
                puntos=puntos_registrados,
                evidencia=evidencia  # Guarda la imagen
            )
        else:
            messages.error(request, "No se subió ningún archivo de evidencia.")

        # Mensaje de éxito
        messages.success(request, f'Los puntos fueron {"sumados" if puntos > 0 else "restados"} correctamente')
        return redirect('points', id_usuario=usuario_id)

    # Si el método no es POST
    messages.error(request, 'Método no permitido')
    return redirect('tables')


"""
    Configuracion de usuarios
"""

@login_required
@rol_required([1]) # Solo puede entrar un administrador
def register(request):    
    form = RegistroForm()
    
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        
        if form.is_valid():
            # Se obtienen los datos del formulario
            numero_documento = form.cleaned_data['numero_documento']
            nombre_usuario = form.cleaned_data['nombre_usuario']
            email = form.cleaned_data['email']
            contrasena = form.cleaned_data['contrasena']
            confirmar_contrasena = form.cleaned_data['confirmar_contrasena']
            
            # Validar contraseñas
            if contrasena != confirmar_contrasena:
                messages.error(request, "Las contraseñas no coinciden.")
                return render(request, "admin_dashboard/register.html", {'form': form})
            
            # Validar si el email ya existe
            if Usuario.objects.filter(email=email).exists():
                messages.error(request, "El correo ya está registrado.")
                return render(request, "admin_dashboard/register.html", {'form': form})
            
            # Validar si el número de documento ya existe
            if Usuario.objects.filter(id=numero_documento).exists():
                messages.error(request, "El número de documento ya está registrado.")
                return render(request, "admin_dashboard/register.html", {'form': form})
            
            # Crear usuario
            new_user = Usuario.objects.create(
                id=numero_documento,
                nombre=nombre_usuario,
                email=email,
                password=make_password(contrasena),  # Encriptar la contraseña
            )
            new_user.save()  # Guardar en la base de datos
            
            return redirect('register')
        else:
            messages.error(request, "El formulario contiene errores. Por favor, corrígelos.")
    
    return render(request, "admin_dashboard/user_sessions/register.html", {'form': form})

def login(request):
    if request.method == 'POST':
        email = request.POST['email']  # Capturar email del formulario
        password = request.POST['password']  # Capturar contraseña del formulario

        try:
            # Intentar obtener al usuario por email
            user = Usuario.objects.get(email=email)
            
            # Verificar si la contraseña ingresada coincide con la almacen
            if check_password(password, user.password):
                # Guardar información del usuario en la sesión
                request.session['user_id'] = user.id
                request.session['user_name'] = user.nombre
                request.session['user_rol'] = user.rol
                
                return redirect('index')  # Redirigir a la página principal
            else:
                messages.error(request, "Contraseña incorrecta. Por favor, inténtalo nuevamente.")
        except Usuario.DoesNotExist:
            messages.error(request, "El correo electrónico no está registrado.")
    return render(request, "admin_dashboard/user_sessions/login.html")

def logout(request):
    # Eliminar la sesion
    request.session.clear()
    return redirect('login')

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        if email:
            try:
                user = Usuario.objects.get(email=email)
                
                if user.rol == 1:
                    password = 'Adm1n@pos-2025'
                else:
                    password = 'F@miliapos-2024'
                
                send_mail(
                    '¡Recuerda tu contraseña!',
                    f'Hola {user.nombre}, tu contraseña asignada a sido la siguiente: {password}',
                    'nicolas.paulo.vega06@gmail.com',
                    [email],
                    fail_silently=False,
                )
                
                messages.success(request, 'Se envio un mensaje a tu correo de manera exitosa!')
                return redirect('forgot-password')
            except Usuario.DoesNotExist:
                messages.error(request, 'El correo no se encuentra registrado')
        else:
            messages.error(request, 'Por favor, ingresa un correo valido.')
    return render(request, 'admin_dashboard/forgot-password.html')

@login_required
@rol_required([1]) # Solo puede entrar un administrador
def user_update(request, id):
    # Buscar usuario
    usuario = get_object_or_404(Usuario, id=id)
    # Cargar formulario POST
    if request.method == 'POST':
        # Cargar formulario con imagenes
        form = UsuarioEditForm(request.POST, request.FILES)
        
        # Validar si el correo es valido
        if form.is_valid():
            # Buscar si hay imagen
            nueva_imagen = form.cleaned_data.get('img')
            # Si no hay entonces cargar una nueva
            if nueva_imagen:
                usuario.img = nueva_imagen
            # Cargar datos a cambiar
            usuario.rol = form.cleaned_data['rol']
            usuario.nombre = form.cleaned_data['nombre']
            usuario.email = form.cleaned_data['email']
            usuario.habilitacion = form.cleaned_data['habilitacion']
            # Guardar formulario
            usuario.save()
            # Mensaje de exito
            messages.success(request, "Usuario actualizado con éxito")
            # Volver al los usuarios
            return redirect('tables')
    else:
        # Cargar formulario con datos
        form = UsuarioEditForm(initial={
            'rol': usuario.rol,
            'img': usuario.img,
            'nombre': usuario.nombre,
            'email': usuario.email,
            'habilitacion': '1' if usuario.habilitacion else '0',
        })
    # Renderizar la plantilla de actualizacion de usuario
    return render(request, 'admin_dashboard/user_sessions/update_user.html', {'form': form, 'usuario': usuario})

@login_required
@rol_required([1])
def evidence(request, id_usuario):
    id_usuario = Usuario.objects.get(id=id_usuario)
    evidence = Puntos.objects.filter(usuario=id_usuario)
    return render(request, 'admin_dashboard/evidence.html', {'evidence': evidence, 'user': id_usuario})

@login_required
@rol_required([1])
def exportar_a_excel(request):
    # Obtener el año y mes actuales
    fecha_actual = datetime.now()
    anio_actual = fecha_actual.year
    mes_actual = fecha_actual.month

    # Definir el primer y último día del mes actual
    primer_dia_mes = date(anio_actual, mes_actual, 1)
    if mes_actual == 12:
        ultimo_dia_mes = date(anio_actual, mes_actual, 31)
    else:
        ultimo_dia_mes = date(anio_actual, mes_actual + 1, 1) - timedelta(days=1)

    # Filtrar los datos entre el primer y último día del mes actual
    datos = Puntos.objects.filter(
        fecha__range=[primer_dia_mes, ultimo_dia_mes]
    ).select_related('usuario', 'subcategoria')

    # Crear el archivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Datos"

    # Encabezados
    encabezados = ["Usuario", "Subcategoría", "Puntos", "Fecha"]
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = sheet.cell(row=1, column=col_num)
        celda.value = encabezado
        celda.font = Font(bold=True)
        celda.alignment = Alignment(horizontal="center")

    # Llenar las filas con los datos
    for fila, punto in enumerate(datos, start=2):
        sheet.cell(row=fila, column=1).value = punto.usuario.nombre
        sheet.cell(row=fila, column=2).value = punto.subcategoria.nombre
        sheet.cell(row=fila, column=3).value = punto.puntos
        sheet.cell(row=fila, column=4).value = punto.fecha.strftime("%Y-%m-%d")

    # Ajustar ancho de las columnas
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 10
    sheet.column_dimensions["D"].width = 15

    # Generar respuesta HTTP con el archivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="reporte_datos_{anio_actual}_{mes_actual}.xlsx"'
    workbook.save(response)
    return response